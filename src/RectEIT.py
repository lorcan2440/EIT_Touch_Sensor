# local imports
import kg_robot as kgr
import waypoints as wp
from utils import *

# third party imports
import numpy as np
import pandas as pd
import serial
import openpyxl
from matplotlib import pyplot as plt
from matplotlib.colors import ListedColormap

# built ins
import logging
import os
import subprocess
import datetime as dt
import random
import time
import math
import pickle


class RectEIT:
    '''
    Represents a combined Electrical Impedance Tomography (EIT) setup consisting of:
    - A Universal Robot 5 kg payload (UR5) robot arm with a claw tool attached (the end effector)
    - The claw tool has two fingers, one fixed and one movable by an Arduino controlled motor
    - A rectangular frame containing a soft sensorised gelatin hydrogel surrounded by 32 electrodes
    - A microcontroller device running an EIT experiment on the electrodes
    - The UR5, Arduino and EIT microcontroller are all connected to this PC running the program

    Our EIT setup uses:
    - 32 electrodes as M4 bolts screwed through the perimeter of a 3D printed rectangular frame
    - Touch region dimensions: 120 x 92 mm
    - Finger dimensions: 2 mm radius, 15 - 85 mm separation
    - Hydrogel composition: wt% 1 : 1.5 : 2.5 : 0.2 : 0.1 - gelatin (pork, 240 bloom, MM Ingredients) 
    : glycerol (Fisher Scientific) : tap water : citric acid monohydrate (Fisher Scientific) 
    : sodium chloride (table salt), incubated in water bath at 65 degrees C for 1 hour before
    casting into a 3D printed mould
    - EIT current input: ??? mA a.c., ? kHz, through opposite electrodes
    '''

    # robot parameters
    FINGER_MAX_SEP = 85 * 0.001  # m
    FINGER_MIN_SEP = 15 * 0.001  # m
    FINGER_RADIUS = 2 * 0.001  # m
    PRESS_DEPTH = 95 * 0.001  # 95 mm
    CORNER_ORIGIN = np.array(wp.home)  # fixed finger 4 cm above skin at corner near electrode 1, other finger parallel to 
    CLAW_TCP = wp.claw_tcp  # tip of fixed finger relative end effector
    PIVOT_R = abs(CLAW_TCP[0])  # 0.055398 m: distance between centre of end effector and fixed finger

    # EIT parameters
    NUM_ELECTRODES = 32
    NUM_VOLTAGES = 1024  # n * (n - 4) for old board, n^2 for new board
    FRAME_X = 0.08  # m
    FRAME_Y = 0.12  # m
    FRAME_DEADZONE = 0.01  # m, distance from edge of frame that fingers cannot reach
    X_MIN, X_MAX = FINGER_RADIUS + FRAME_DEADZONE, FRAME_X - FINGER_RADIUS - FRAME_DEADZONE
    Y_MIN, Y_MAX = FINGER_RADIUS + FRAME_DEADZONE, FRAME_Y - FINGER_RADIUS - FRAME_DEADZONE
    GRID_DIV_X = 6  # resolution of output grid in x direction
    GRID_DIV_Y = 8  # resolution of output grid in y direction
    ELECTRODE_TO_EDGE_DIST = 0.006  # m, distance from edge of frame to centre of electrode

    # communications
    ARDUINO_COMMS = {'port': 'COM8', 'baudrate': 115200}
    ROBOT_COMMS = {'port': 30010, 'db_host': '169.254.5.1'}
    EIT_COMMS = {'port': 'COM9', 'baudrate': 9600}  # timeout: 5

    # file paths
    _OUTPUT_FILE = '../output/EIT_Data_Gelatin_2_fingers_4_dof.xlsx'
    _OUTPUT_BASELINE = '../output/EIT_Baselines.xlsx'
    _LOG_FILE = '../output/EIT_log.txt'
    _NOTIFIER_SCRIPT = 'exp_complete_notifier.ps1'

    # misc
    TRIALS_PER_TEMP_WB = 100

    def __init__(self):
        self.VOLTAGES_HEADERS = ['v' + str(i) for i in range(self.NUM_VOLTAGES)]
        self.POSITIONS_HEADERS = ['x0', 'x1', 'y0', 'y1']
        self.init_logger()
    
    def init_logger(self):
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler = logging.FileHandler(self._LOG_FILE)
        file_handler.setFormatter(file_formatter)
        stream_handler = logging.StreamHandler()
        stream_handler.setFormatter(ColouredLoggingFormatter())
        self.logger.addHandler(file_handler)  # send logs to file
        self.logger.addHandler(stream_handler)  # send logs to console

    def connect_to_hardware(self, move: bool = True, confirm: bool = True) -> None:
        '''
        Connects to the robot (via Ethernet), the Arduino (via USB)
        and the EIT microcontroller (via USB).
        
        ### Arguments
        #### Required
        #### Optional
        - `move` (bool, default = True): if true, move the robot to its starting position.
        - `confirm` (bool, default = True): if true, ask the user to confirm before moving the robot.
        '''   
        self.robot = kgr.kg_robot(**self.ROBOT_COMMS)
        self.arduino = serial.Serial(**self.ARDUINO_COMMS)
        self.eit = serial.Serial(**self.EIT_COMMS)
        time.sleep(0.5)
        self.eit.reset_input_buffer()
        self.eit.reset_output_buffer()
        self.eit.flush()
        self.eit.read_all()
        time.sleep(0.5)
        self.robot.set_tcp(self.CLAW_TCP)
        self.logger.info(f'Robot initial position: {self.robot.getl()}')
        self.logger.info(f'Robot initial joint angles: {self.robot.getj()}')
        if move:
            if confirm:
                print(f'The robot is about to move to the position {self.CORNER_ORIGIN}.\n'
                    'Ensure the area is clear of obstacles before continuing.')
                ans = input('Press Enter to move. To exit the program and NOT move, type "STOP" and press Enter. ')
                if ans.strip().upper() == 'STOP':
                    self.logger.info('The program is closing because "STOP" was entered.')
                    exit()  # abort program
            self.logger.warning('The robot is moving to its starting position.')
            self.move_up_if_too_low()
            self.robot.movel(self.CORNER_ORIGIN, acc=0.01, vel=0.01)
    
    def test_corner_calibration_positions(self) -> None:
        '''
        Moves the robot to the four corners of the frame to test the coordinate system calibration.
        '''
        self.logger.info('Testing corner calibration positions. The robot will move to the four corners of the frame.')
        self.logger.warning('The robot is about to move to the corner near Electrode 1 (origin).')
        input('Press Enter.')
        self.robot.translatel(self.CORNER_ORIGIN, acc=0.1, vel=0.1)

        self.logger.warning('The robot is about to move to the corner near Electrode 10 (y-axis limit).')
        input('Press Enter.')
        self.robot.translatel_rel([0, self.FRAME_Y, 0], acc=0.1, vel=0.1)

        self.logger.warning('The robot is about to move to the corner near Electrode 26 (x-axis limit).')
        input('Press Enter.')
        self.robot.translatel_rel([self.FRAME_X, -self.FRAME_Y, 0], acc=0.1, vel=0.1)

        self.logger.warning('The robot is about to move to the corner near Electrode 17 (opposite origin).')
        input('Press Enter.')
        self.robot.translatel_rel([0, self.FRAME_Y, 0], acc=0.1, vel=0.1)

        self.logger.warning('The robot is about to return to the origin.')
        input('Press Enter.')
        self.robot.translatel_rel([-self.FRAME_X, -self.FRAME_Y, 0], acc=0.1, vel=0.1)

        home_after = self.robot.getl()
        if np.allclose(home_after, self.CORNER_ORIGIN, atol=0.001):  # 1 mm deviation
            self.logger.info('Corner calibration positions test complete. '
                'Current position matches origin to within 1 mm.')
        else:
            self.logger.error(f'Corner calibration positions test failed. '
                f'Expected to be at or near {self.CORNER_ORIGIN} but actually at {home_after}.')
    
    def test_gear_motor(self, num_times: int = 1):
        '''
        Performs a simple Arduino gear motor test, checking that it does not get stuck at one end.
        
        ### Arguments
        #### Required
        #### Optional
        - `num_times` (int, default = 1): number of times to open and close the claw.
        '''
        self.logger.info('Testing gear motor: close -> open -> close -> up -> down.')
        self.move_gear(self.FINGER_MIN_SEP)
        if num_times is None:
            while True:
                self.test_gear_motor(num_times=1)
        for _ in range(num_times):
            self.move_gear(self.FINGER_MAX_SEP)  # open
            self.move_gear(self.FINGER_MIN_SEP)  # close
            self.move_gear(self.FINGER_MIN_SEP, vertical=0)  # up
            self.move_gear(self.FINGER_MIN_SEP, vertical=1)  # down
            time.sleep(0.5)
        self.logger.info('Motor test complete.')
    
    def move_up_if_too_low(self):
        if (tmp_z := self.robot.getl()[2]) <= 0.18:
            self.logger.warning('Robot is too low. Moving up.')
            self.robot.translatel_rel([0, 0, (self.CORNER_ORIGIN[2] - tmp_z), 0, 0, 0], acc=0.05, vel=0.05)
    
    def read_eit_voltages(self, as_array: bool = True, check_full: bool = True) -> list | np.ndarray:
        '''
        Measures the EIT voltages and returns the results.
        Unused combinations (the four adjacent electrodes) are set to zero,
        so the dimension is n^2.
        
        ### Arguments
        #### Required
        #### Optional
        - `as_array` (bool, default = True): if True, return the results as a numpy array, else as a list.
        - `check_full` (bool, default = True): if True, check that the received data is complete.
        Set this option to false if testing by making the EIT microcontroller send only a subset of the data.

        ### Returns
        - `list | np.ndarray`: voltages, in volts. If array, shape is (1024,). 
        '''
        raw_data = self.eit.read_until(expected=b'\n')
        raw_data = raw_data.decode('utf-8').strip().split(',')
        #raw_data = self.eit.readline().decode('utf-8').strip().split(',')
        self.eit.reset_input_buffer()
        self.eit.flush()
        raw_data = [float(v) for v in raw_data[:-1]]
        if len(raw_data) != 1024 and check_full:
            self.logger.warning(f'Received an incomplete set of voltages: size {len(raw_data)}, '
                f'expected 1024. Retrying.')
            time.sleep(0.1)
            return self.read_eit_voltages(as_array=as_array)
        else:
            return raw_data if not as_array else np.array(raw_data)
    
    def move_gear(self, dist: float, vertical: int = 1) -> None:
        '''
        Instructs the Arduino to move the gear motor in order to open/close the claw.
        The speed of the gear is fixed and is determined by the power supply voltage.
        
        ### Arguments
        #### Required
        - `dist` (float): desired separation between claws, in metres.
        - `vertical` (int, default = 1): if 1, move the fixed finger down.
        If 0, move the fixed finger up.
        #### Optional
        
        ### Raises
        - `ValueError`: if the desired separation is outside the allowable range.
        '''
        if not self.FINGER_MIN_SEP <= dist <= self.FINGER_MAX_SEP:
            raise ValueError(f'The specified finger positions are out of the allowable range.')
        if vertical not in [0, 1]:
            raise ValueError(f'Vertical must be 0 or 1.')
        
        # data sent as a byte b'XYYYYYYY', where X is vertical and YYYYYYY is servo angle
        servo_angle = gear_metres_to_servo(dist)
        servo_angle_bin = bin(servo_angle)[2:].zfill(7)
        vertical_bin = bin(vertical)[2:].zfill(1)
        data_bin = vertical_bin + servo_angle_bin
        data = bytes([int(data_bin, 2)])
        self.logger.info(f'Gear(s) moving: dist {round(dist * 1000, 2)} mm, {"down" if vertical == 1 else "up"}. '
            f'Sending serial byte {data}')
        self.arduino.write(data)
        time.sleep(1.5)

    def disconnect_from_hardware(self) -> None:
        '''
        Closes all serial connections.
        '''        
        self.logger.info('Closing all serial connections.')
        self.robot.close()
        self.arduino.close()
        self.eit.close()

    @staticmethod
    def show_voltage_map(voltages: np.ndarray, baseline: np.ndarray = None):
        '''
        Plots a colour map of the voltages, represented as a n by n array.
        Can optionally subtract a baseline from the voltages.
        If `voltages` contains multiple rows, their average is plotted.

        The pixel at position [i, j] (indices counting from 1 upto 32) represents
        the voltage between electrodes j and j + 1 (mod 32) when current is injected
        through electrode i and drained through ground electrode i + 1 (mod 32).
        This pattern is determined by the 'adjacent' setting `AD` in the EIT microcontroller
        (`main.cpp`, call to `read_frame(...)`, line 148).
        
        ### Arguments
        #### Required
        - `voltages` (np.ndarray): array of voltages, in volts. Shape: (num_trials, 1024) or (1024,)
        #### Optional
        - `baseline` (np.ndarray, default = None): array of baselines, in volts. Shape: (1024,)
        ''' 
        if voltages.ndim > 1:
            voltages = np.mean(voltages, axis=0)
        if baseline is not None:
            voltages -= baseline
        size_1d = voltages.shape[0]
        size_width = round(2 + np.sqrt(4 + size_1d))
        size_height = round(size_width - 4)
        plt.title(f'{"Relative " if baseline is not None else ""}Voltage map (averaged over trials)')
        plt.imshow(voltages.reshape(size_width, size_height), cmap='viridis')
        plt.colorbar()
        plt.show()
    
    def show_voltage_graph(self, voltages: np.ndarray, baseline: np.ndarray = None):
        '''
        Plots a line graph of the voltages.
        Expect periodic spikes of period 32.
        If a baseline is provided, the graph will also show the voltages relative to the baseline,
        as well as the baseline itself.
        
        ### Arguments
        #### Required
        - `voltages` (np.ndarray): array of voltages, in volts. Shape: (num_trials, 1024) or (1024,)
        #### Optional
        - `baseline` (np.ndarray, default = None): array of baselines, in volts. Shape: (1024,)
        '''
        if baseline is None:
            plt.title('Touching voltages (raw values)')
            plt.xlabel('Reading number')
            plt.ylabel('Voltage / $ V $')
            for i, row in enumerate(voltages, start=1):
                plt.plot(list(range(self.NUM_VOLTAGES)), row, alpha=0.5)
            plt.plot(list(range(self.NUM_VOLTAGES)), np.mean(voltages, axis=0), color='black', label='Average')
            plt.legend(loc='upper right')
        else:
            baseline = np.tile(baseline, (voltages.shape[0], 1))
            fig, (ax_b, ax_raw, ax_rel) = plt.subplots(1, 3, figsize=(12, 4))
            fig.suptitle('EIT Data Results: Voltages')
            fig.tight_layout(pad=3)
            ax_b.set_title('Baseline voltages, $ V_0 $')
            ax_b.set_xlabel('Reading number')
            ax_b.set_ylabel('Voltage / $ V $')
            ax_b.plot(list(range(self.NUM_VOLTAGES)), baseline[0], color='black', label='Baseline')
            ax_raw.set_title('Touching voltages (raw values, $ V $)')
            ax_raw.set_xlabel('Reading number')
            ax_raw.set_ylabel('Voltage / $ V $')
            for i, row in enumerate(voltages, start=1):
                ax_raw.plot(list(range(self.NUM_VOLTAGES)), row, label=f'Trial {i}')
            ax_raw.legend(loc='upper right')
            ax_rel.set_title('Touching voltages (relative to baseline, $ \Delta V = V - V_0 $)')
            ax_rel.set_xlabel('Reading number')
            ax_rel.set_ylabel('Relative Voltage / $ V $')
            for i, row in enumerate(voltages - baseline, start=1):
                ax_rel.plot(list(range(self.NUM_VOLTAGES)), row, label=f'Trial {i}', alpha=0.5)
            ax_rel.plot(list(range(self.NUM_VOLTAGES)), np.mean(voltages - baseline, axis=0), color='black', label='Average')
            ax_rel.legend(loc='upper right')
        plt.show()

    def get_baseline(self, output_file: str = None, average: bool = False,
            num_samples: int = 10, **kwargs) -> np.ndarray:
        '''
        Collects a number of EIT readings with the fingers not touching the skin (baseline).
        If `average` is True, the results will be averaged.
        
        ### Arguments
        #### Required
        #### Optional
        - `output_file` (str, default = None): if provided, save the baseline to this file path.
        - `average` (bool, default = False): if True, average the results.
        - `num_samples` (int, default = 10): number of readings to take (to be averaged).
        
        ### Returns
        - `np.ndarray`: array of baseline voltages, in volts. Shape: (num_trials, 1024) or (1024,).
        '''
        self.logger.info('Obtaining baseline EIT data. The robot will not move.')
        dataset = []
        for sample in range(num_samples):
            dataset.append(self.read_eit_voltages(**kwargs))
        self.logger.info('Baseline EIT data obtained.')
        if average:
            if output_file is not None:
                self.save_to_excel([np.mean(dataset, axis=0)], ['Baseline'], [self.VOLTAGES_HEADERS], output_file, **kwargs)
            return np.mean(dataset, axis=0)  # shape: (1024,)
        else:
            if output_file is not None:
                self.save_to_excel([np.array(dataset)], ['Baseline'], [self.VOLTAGES_HEADERS], output_file, **kwargs)
            return np.array(dataset)  # shape: (num_trials, 1024)
    
    def convert_xy_to_disp_map(self, positions: list | np.ndarray) -> np.ndarray:
        '''
        Given a set of frame coordinates, return an image-like array showing
        the position of the fingers. It is OK for some of the values to be None,
        which represent empty positions.
            
        ### Arguments
        #### Required
        - `positions` (list | np.ndarray): list of position coordinates
        [x0, x1, ..., y0, y1, ...], in metres in the frame coordinates.
        Shape: (num_trials, 2 * num_fingers) or can be (2 * num_fingers,) if only one.
        #### Optional
            
        ### Returns
        - `np.ndarray`: array of images showing the finger positions. Each array
        will be 1 if there is a press within that tile, and 0 otherwise.
        The radius of the finger is neglected.
        Shape: (num_trials, self.GRID_DIV_X, self.GRID_DIV_Y) or
        (self.GRID_DIV_X, self.GRID_DIV_Y) if only one.
        '''

        tile_x = self.FRAME_X / self.GRID_DIV_X
        tile_y = self.FRAME_Y / self.GRID_DIV_Y

        if isinstance(positions, list):
            positions = np.array(positions)
        if positions.ndim == 1:
            positions = positions.reshape(1, -1)

        disp_maps = []
        for pos in positions:
            pos = pos[~np.isnan(pos)]  # remove None entries, preserve order
            disp_map = np.zeros((self.GRID_DIV_X, self.GRID_DIV_Y))
            x_list = pos[::len(pos) // 2]
            y_list = pos[len(pos) // 2::]
            for x, y in zip(x_list, y_list):
                disp_map[int(x // tile_x), int(y // tile_y)] = 1.0
            disp_maps.append(disp_map)
        return np.array(disp_maps) if len(disp_maps) > 1 else disp_maps[0]

    def show_disp_map(self, img: np.ndarray) -> None:
        '''
        Show an illustration of the finger positions as an array of pixels.
        
        ### Arguments
        #### Required
        - `img` (np.ndarray): an pixel array where each pixel is 1 for a press and 0 if not.
        Shape: (self.GRID_DIV_X, self.GRID_DIV_Y).
        #### Optional
        '''        
        cMap = ListedColormap(['white', 'green'])
        fig, ax = plt.subplots()
        heatmap = ax.pcolor(img, cmap=cMap)
        cbar = plt.colorbar(heatmap, aspect=2, shrink=0.2)
        cbar.ax.get_yaxis().set_ticks([])
        for j, lab in enumerate(['No', 'Yes']):
            cbar.ax.text(0.5, (2 * j + 1) / 4.0, lab, ha='center', va='center')
        cbar.ax.get_yaxis().labelpad = 15
        cbar.ax.set_ylabel('Contact', rotation=270)
        ax.set_xlabel('x position (tile number)')
        ax.set_ylabel('y position (tile number)')
        ax.set_xticks(np.arange(0, self.GRID_DIV_X + 1, 1))
        ax.set_yticks(np.arange(0, self.GRID_DIV_Y + 1, 1))
        ax.set_xlim(0, self.GRID_DIV_X)
        ax.set_ylim(0, self.GRID_DIV_Y)
        ax.set_title('Finger position')
        ax.grid(linewidth=2, visible=True, which='major', alpha=1.0)
        ax.grid(which='minor', alpha=0.0)
        plt.show()

    def validate_positions(self, positions: list | np.ndarray, raise_error: bool = True) -> bool:
        '''
        Checks whether a given position of the fingers is valid.
        
        ### Arguments
        #### Required
        - `positions` (list | np.ndarray): coordinates [x0, x1, y0, y1] (0: fixed finger, 1: moving finger), in m.
        #### Optional
        - `raise_error` (bool, default = True): if True, raise a ValueError if the positions are invalid.
        
        ### Returns
        - `bool`: True if the positions are valid, else False.
        
        ### Raises
        - `ValueError`: if the positions are invalid and `raise_error` is True.
        '''

        positions = list(positions)
        x0, x1, y0, y1 = positions
        num_fingers = 2
        if None in positions and positions.count(None) == 2:
            num_fingers = 1
            x0 = x1 if x0 is None else x0
            y0 = y1 if y0 is None else y0
            x1 = x0 if x1 is None else x1
            y1 = y0 if y1 is None else y1

        # check if finger x is too small or too large
        if not (self.X_MIN <= x0 <= self.X_MAX and self.X_MIN <= x1 <= self.X_MAX):
            if raise_error:
                raise ValueError(f'Finger x position out of range. Got {x0}, {x1}, must be in range '
                    f'[{self.X_MIN}, {self.X_MAX}].')
            else:
                return False

        # check if finger y is too small or too large
        if not (self.Y_MIN <= y0 <= self.Y_MAX and self.Y_MIN <= y1 <= self.Y_MAX):
            if raise_error:
                raise ValueError(f'Finger y position out of range. Got {y0}, {y1}, must be in range '
                    f'[{self.Y_MIN}, {self.Y_MAX}].')
            else:
                return False

        # check distance between fingers
        dist = np.hypot(x0 - x1, y0 - y1)
        if not self.FINGER_MIN_SEP <= dist <= self.FINGER_MAX_SEP and num_fingers == 2:
            if raise_error:
                raise ValueError(f'Fingers too close or too far apart. Got {dist}, '
                    f'must be in range [{self.FINGER_MIN_SEP}, {self.FINGER_MAX_SEP}].')
            else:
                return False

        # check angle between fingers
        angle = np.arctan2(x0 - x1, y0 - y1)
        if not -2.4 <= angle <= 0.75:
            if raise_error:
                raise ValueError(f'Fingers at invalid angle. Got {angle}, '
                    f'must be in range [-2.4, 0.75].')
            else:
                return False

        return True

    def frame_xy_to_robot_xy(self, x: float, y: float, _z: float) -> np.ndarray:
        '''
        Converts a given (x, y) in the frame coordinates to an absolute (x, y) in the robot coordinates.
        The z coordinate is provided for compatibility and is maintained.
        NOTE: The angles are all set to zero so this must only be used for translation, not movements.
        Using this on a movement will likely crash the robot!

        Example usage: 
        
        `self.robot.translatel(self.frame_xy_to_robot_xy(x, y, z))`, or
        `self.robot.translatel_rel(self.frame_xy_to_robot_xy(dx, dy, dz))`
        
        ### Arguments
        #### Required
        - `x` (float): x-coordinate in frame [mm], positive directed towards Electode 26 from origin
        - `y` (float): y-coordinate in frame [mm], positive directed towards Electode 10 from origin
        - `_z` (float): z-coordinate in frame [mm], positive is vertically upwards
        #### Optional
        
        ### Returns
        - `np.ndarray`: array for the position pose of the robot, in metres. Shape: (6,), last three are zeros
        '''    
        robot_x = self.CORNER_ORIGIN[0] + x
        robot_y = self.CORNER_ORIGIN[1] + y
        return np.array([robot_x, robot_y, _z, 0, 0, 0])

    def save_to_excel(self, data_list: list[np.ndarray], sheet_names_list: list[str],
            header_row_list: list[list[str]], filename: str, **kwargs):
        '''
        Append data arrays to Excel sheets row-wise. If the file or a sheet does not exist,
        it will be created. Data will be appended with headers if the sheet is new.

        Example usage:
        ```
        data1 = np.array([[1, 5, 10], [2, 4, 11]])
        data2 = np.array([[8, 4], [10, 0.5]])
        sheet_names = ['First', 'Second']
        header_rows = [['Alpha', 'Beta', 'Gamma'], ['Delta', 'Epsilon']]
        self.save_to_excel([data1, data2], sheet_names, header_rows, 'test.xlsx')
        ```
        
        ### Arguments
        #### Required
        - `data_list` (list[np.ndarray]): a list of 2D NumPy arrays to save to each sheet, row-wise.
        - `sheet_names_list` (list[str]): a corresponding list of sheet names to save these arrays in.
        - `header_row_list` (list[list[str]]): a corresponding list of header rows to add to the top.
        - `filename` (str): the output filename.
        #### Optional
        - `trial` (int, default = None): if provided, the function will temporarily save to smaller
        files every `self.TRIALS_PER_TEMP_WB` trials in order to prevent slowdown with large files.
        These temporary files will all be merged into the target output file (without overwriting) at the end.
        - `force_same_dim` (bool, default = True): if True, raise an ValueError if the arrays
        have mismatching numbers of rows. 
        
        ### Raises
        - `ValueError`: if `force_same_dim` is True and any arrays have mismatching numbers of rows.
        '''
        force_same_dim = kwargs.get('force_same_dim', True)
        trial = kwargs.get('trial', None)

        if len(data_list) != len(sheet_names_list):
            raise ValueError('The number of data arrays and sheet names do not match.')
        
        # if there are any 1D arrays, they will be converted to 2D arrays of first dim 1
        for i, data in enumerate(data_list):
            if data.ndim == 1:
                data = data.reshape(1, -1)
                data_list[i] = data

        # check all arrays have same first dim if needed
        if force_same_dim:
            if not all([data.shape[0] == data_list[0].shape[0] for data in data_list]):
                raise ValueError('Arrays have mismatching number of rows. Not saving.')
        
        # set the name of the workbook to save this data to
        if trial is not None:
            this_wb_num = math.ceil(trial / self.TRIALS_PER_TEMP_WB)
            this_wb_name = os.path.splitext(filename)[0] + f'_Set_{this_wb_num}.xlsx'
        else:
            this_wb_name = filename

        if os.path.exists(this_wb_name):
            # iterate through each sheet and add the data
            for i, (data, sheet_name, header_row) in enumerate(zip(data_list, sheet_names_list, header_row_list)):
                data_df = pd.DataFrame(data, columns=header_row)
                with pd.ExcelWriter(this_wb_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    sheet_exists = (sheet_name in writer.sheets)
                    start_i = writer.sheets[sheet_name].max_row if sheet_exists else 1
                    self.logger.info(f'Appending data in {sheet_name} ({this_wb_name}) from rows {start_i + 1} to {start_i + data.shape[0]}.')
                    data_df.to_excel(writer, sheet_name=sheet_name, index=False,
                        startrow=start_i, header=(not sheet_exists))
        else:
            # create new workbook
            workbook = openpyxl.Workbook()
            workbook.save(this_wb_name)
            for sheet_name, header_row in zip(sheet_names_list, header_row_list):
                workbook.create_sheet(title=sheet_name)
                with pd.ExcelWriter(this_wb_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    if header_row is not None:
                        pd.DataFrame(columns=header_row).to_excel(writer, sheet_name=sheet_name, index=False)
            self.logger.info(f'Created new Excel file {this_wb_name}.')
            # add data
            _args = [data_list, sheet_names_list, header_row_list, this_wb_name]
            kwargs.pop('trial', None)
            self.save_to_excel(*_args, **kwargs)
        
        if trial is not None and trial == self.num_expts_to_do:
            # this is the last one
            # merge all generated sheets into the main output file
            self.logger.info(f'Merging file fragments into {filename}.')
            from utils import merge_excel_files
            search_str = os.path.splitext(os.path.basename(filename))[0] + '_Set_'
            files_dir = os.path.dirname(filename)
            files_to_merge = [os.path.join(files_dir, f) for f in os.listdir(files_dir) \
                if os.path.basename(f).startswith(search_str)]
            merge_excel_files(files_to_merge, filename)
            # delete the temporary files
            self.logger.info(f'Deleting file fragments.')
            for f in files_to_merge:
                os.remove(f)
            
    def get_voltages_at_rect_pos(self, x0: float, x1: float, y0: float, y1: float, 
            num_samples: int = 10, baseline: np.ndarray = None,
            append_to_dataset: bool = True, average: bool = True,
            **kwargs) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
        '''
        Conduct EIT experiments at a given position in frame coordinates.
        Save and return the results. If (x0, y0) or (x1, y1) are None, only one finger will be used.
        
        ### Arguments
        #### Required
        - `x0` (float): x coordinate of the fixed (vertically moving) finger, in metres
        - `x1` (float): x coordinate of the moving (horizontally moving) finger, in metres
        - `y0` (float): y coordinate of the fixed (vertically moving) finger, in metres
        - `y1` (float): y coordinate of the moving (horizontally moving) finger, in metres
        #### Optional
        - `num_samples` (int, default = 10): number of times to repeat the experiment at this position
        - `baseline` (np.ndarray, default = None): if provided, additionally calculates
        the voltages relative to this baseline
        - `append_to_dataset` (bool, default = True): if True, add the data to `self._OUTPUT_FILE`.
        - `average` (bool, default = True): if True, average the results of each trial.
        
        ### Returns
        - `tuple[np.ndarray, np.ndarray, np.ndarray]`: the positions (x0, x1, y0, y1), absolute voltages,
        and relative voltages, respectively, all as 2D arrays (even if averaged). Relative voltages
        will be None if `baseline` is not provided.
        
        ### Raises
        - `ValueError`: if the position given is invalid. Note that the function will attempt to correct
        the positions once by switching the order, but if this also fails, then this error will be raised.
        '''        
        # set default movement kwargs - planar, vertical, rotational
        trial = kwargs.get('trial', None)
        vel = kwargs.get('vel', [0.1, 0.1, 1.0])
        acc = kwargs.get('acc', [0.1, 0.1, 1.0])
        _warning_count = kwargs.get('_warning_count', 0)

        # check positions are possible
        try:
            num_fingers = 1 if [x0, x1, y0, y1].count(None) == 2 else 2
            positions = np.array([x0, x1, y0, y1])
            self.validate_positions(positions)
        except ValueError:
            if _warning_count >= 1:
                self.logger.error('Both orders failed to produce a valid position. Aborting.')
                raise ValueError('Both orders failed to produce a valid position. Aborting.')
            self.logger.warning('Position is invalid. Trying again with swapped order of fingers.')
            return self.get_voltages_at_rect_pos(x1, x0, y1, y0, num_samples=num_samples, baseline=baseline, 
                append_to_dataset=append_to_dataset, average=average,
                **(kwargs | {'_warning_count': _warning_count + 1}))

        # calculate (r, theta)
        if num_fingers == 2:
            dist = np.hypot(x0 - x1, y0 - y1)  # distance between fingers
            theta = np.arctan2(x0 - x1, y0 - y1)  # clockwise from positive y-axis
        
        # check if robot is too low - if so, move up
        self.move_up_if_too_low()

        # move to position
        if num_fingers == 2:
            self.logger.info(f'Moving to position: '
                f'(x0, x1, y0, y1) = ({round(1000 * x0, 2)}, {round(1000 * x1, 2)}, '
                f'{round(1000 * y0, 2)}, {round(1000 * y1, 2)}); d = {round(1000 * dist, 2)} [mm].')
            x_pivot = x0 - self.PIVOT_R * np.sin(theta)
            y_pivot = y0 + self.PIVOT_R * (1 - np.cos(theta))
            target_xyz = self.frame_xy_to_robot_xy(x_pivot, y_pivot, self.CORNER_ORIGIN[2])
            self.robot.translatel(target_xyz, acc=acc[0], vel=vel[0])
            # rotate to position
            start_joints = self.robot.getj()
            self.robot.movej_rel([0, 0, 0, 0, 0, theta], acc=acc[2], vel=vel[2])
            # extend gear
            self.move_gear(dist)
        elif num_fingers == 1:
            self.logger.info(f'Moving to position: '
                f'(x, y) = ({round(1000 * x1, 2)}, {round(1000 * y1, 2)}).')
            self.move_gear(self.PIVOT_R, vertical=0)
            target_xyz = self.frame_xy_to_robot_xy(x1, y1 + self.PIVOT_R, self.CORNER_ORIGIN[2])
            self.robot.translatel(target_xyz, acc=acc[0], vel=vel[0])
    
        # get touching voltages
        touching_voltages = []
        for sample in range(1, num_samples + 1):
            # move down
            self.robot.translatel_rel([0, 0, -1 * self.PRESS_DEPTH, 0, 0, 0], acc=acc[1], vel=vel[1])
            # take readings
            data = self.read_eit_voltages()
            touching_voltages.append(data)
            self.logger.info(f'Received EIT data (sample {sample}, got {len(data)} values).')
            # move up
            self.robot.translatel_rel([0, 0, self.PRESS_DEPTH, 0, 0, 0], acc=acc[1], vel=vel[1])
        
        # undo rotations (don't undo finger/gear movements)
        if num_fingers == 2:
            self.robot.movej(start_joints, acc=acc[2], vel=vel[2])
        
        # process data
        touching_voltages = np.array(touching_voltages)  # shape: (num_trials, -1)
        if average:
            touching_voltages = np.mean(touching_voltages, axis=0).reshape(1, -1)
            positions = positions.reshape(1, -1)
        else:
            positions = np.tile(positions, (num_samples, 1))
        if baseline is not None:
            baseline = np.tile(baseline, (num_samples, 1))
            rel_voltages = touching_voltages - baseline
        else:
            rel_voltages = None
        if append_to_dataset:
            self.save_to_excel([rel_voltages, positions], ['Voltages', 'Positions'],
                [self.VOLTAGES_HEADERS, self.POSITIONS_HEADERS], self._OUTPUT_FILE, trial=trial)

        return positions, touching_voltages, rel_voltages
    
    def get_random_pos(self, num_fingers: int = 2) -> np.ndarray:
        '''
        Generate a random valid position of the fingers.

        ### Arguments
        #### Required
        #### Optional
        - `num_fingers` (int, default = 2): whether to get a position with 1 or 2 fingers

        ### Returns
        - `np.ndarray`: random valid [x0, x1, y0, y1] in frame coordinates.
        If `num_fingers` is 1, the values for x0 and y0 will be None.
        '''        
        x0, x1, y0, y1 = 0, 0, 0, 0
        # generate random coords and check until valid
        while not self.validate_positions([x0, x1, y0, y1], raise_error=False):
            x0 = np.random.uniform(self.X_MIN, self.X_MAX) if num_fingers == 2 else None
            y0 = np.random.uniform(self.Y_MIN, self.Y_MAX) if num_fingers == 2 else None
            x1 = np.random.uniform(self.X_MIN, self.X_MAX)
            y1 = np.random.uniform(self.Y_MIN, self.Y_MAX)
            r = np.hypot(x0 - x1, y0 - y1) if num_fingers == 2 else None
            theta = np.arctan2(x0 - x1, y0 - y1) if num_fingers == 2 else None
            #self.logger.info(f'(x0, y0) = ({round(x0, 4)}, {round(y0, 4)}), (x1, y1) = ({round(x1, 4)}, {round(y1, 4)}), (r, theta) = ({round(r, 4)}, {round(theta, 4)})')
        return np.array([x0, x1, y0, y1])

    def get_randomised_data(self, num_trials: int = 100, append_to_dataset: bool = True,
            show_each: bool = False, baseline: bool = False, take_baseline_every: int = 10,
            num_fingers: int = 2):
        '''
        Conduct several EIT experiments at randomised positions. Save (but do not return) the results.
        In each experiment, only one sample will be collected.
        
        ### Arguments
        #### Required
        #### Optional
        - `num_trials` (int, default = 100): number of data points to collect
        - `append_to_dataset` (bool, default = True): if true, update the Excel file with the new data
        - `show_each` (bool, default = False): if true, show the voltage graph after each experiment
        - `baseline` (bool, default = False): if true, take a baseline before every trial
        - `take_baseline_every` (int, default = 10): if `baseline` is true, take a baseline every
        `take_baseline_every` trials
        - `num_fingers` (int, default = 2): whether to use 1 or 2 fingers in the experiments
        '''

        # start a counter to keep track of how many have been done
        self.num_expts_to_do = num_trials

        times_per_trial = []
        for trial in range(1, num_trials + 1):
            self.logger.warning(f'------ Start of trial {trial} / {num_trials} ------ ')
            time_start = time.time()
            # get baseline
            if baseline and (trial - 1) % take_baseline_every == 0:
                baseline_data = self.get_baseline(output_file=self._OUTPUT_BASELINE, num_samples=1).reshape((-1,))
            # get a random position
            x0, x1, y0, y1 = self.get_random_pos(num_fingers=num_fingers)
            # conduct one experiment and save
            position, abs_voltage, rel_voltage = self.get_voltages_at_rect_pos(
                x0, x1, y0, y1, num_samples=1, baseline=baseline_data, append_to_dataset=append_to_dataset, trial=trial)
            # calculate expected time to finish all trials
            times_per_trial.append(time.time() - time_start)
            time_remaining = round(np.mean(times_per_trial) * (num_trials - trial), 2)
            time_remaining = seconds_to_hms(time_remaining)
            self.logger.info(f'Trial {trial} / {num_trials} done. '
                             f'Expected time remaining: {time_remaining}.')
            if show_each:
                self.show_voltage_graph(abs_voltage, baseline=baseline_data)
        
        # notify that experiment is complete
        total_time_taken = seconds_to_hms(sum(times_per_trial))
        self.logger.info(f'Randomised data collection complete. {num_trials} data points obtained.')
        send_notification_alarm('EIT Experiment Completed!',
            f'{num_trials} points added in {total_time_taken}. Click to open the Excel workbook.',
            self._OUTPUT_FILE, self._NOTIFIER_SCRIPT, self.logger)
