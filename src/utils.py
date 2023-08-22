import logging
import pandas as pd
import numpy as np
import math
import openpyxl
import os
import datetime as dt
import subprocess

class ColouredLoggingFormatter(logging.Formatter):

    grey = "\x1b[38;20m"
    yellow = "\x1b[33;20m"
    red = "\x1b[31;20m"
    bold_red = "\x1b[31;1m"
    reset = "\x1b[0m"
    format = '%(levelname)s: %(message)s'

    FORMATS = {
        logging.DEBUG: grey + format + reset,
        logging.INFO: grey + format + reset,
        logging.WARNING: yellow + format + reset,
        logging.ERROR: red + format + reset,
        logging.CRITICAL: bold_red + format + reset
    }

    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)

def gear_metres_to_servo(dist: float) -> int:
    '''
    Convert a desired separation between fingers to a servo setting for the claw.
    
    ### Arguments
    #### Required
    - `dist` (float): centre-to-centre separation distance of fingers, in metres.
    #### Optional
    
    ### Returns
    - `int`: angle to turn the motor, in degrees.
    '''
    dist_mm = dist * 1000
    return int(math.degrees((dist_mm - 15) / 45))


def send_notification_alarm(title: str, message: str, callback_filename: str,
        script_path: str, logger: logging.Logger):
    '''
    Sends a Windows desktop notification (alarm) with a clickable callback to open a file.
    Uses `exp_complete_notifier.ps1` [self._NOTIFIER_SCRIPT] PowerShell script.
        
    #### Arguments
        
    `title` (str): notification title text.
    `message` (str): notification body text.
    `callback_filename` (str): file path to open when clicked, relative to cwd.
    '''
    # get log file from logger
    log_file = logger.handlers[0].baseFilename

    with open(log_file, 'a') as f:
        try:
            command = f'powershell.exe -ExecutionPolicy Bypass -File {script_path} ' + \
                f'"{title}" "{message}" "{callback_filename}"'
            subprocess.check_call(command, shell=True, stderr=f, stdout=f)
        except subprocess.CalledProcessError as e:
            logger.warning(f'Failed to send notification. Error: {e.returncode} - {e}')


def seconds_to_hms(seconds: float) -> str:
        '''
        Utility to convert a number of seconds to a string in the format 'HHhMMmSSs'.
        Used to report the time taken by the EIT experiments.
        
        ### Arguments
        #### Required
        - `seconds` (float): number of seconds.
        #### Optional
        
        ### Returns
        - `str`: string in the format 'HHhMMmSSs'.
        '''
        result = dt.datetime.strptime(
            str(dt.timedelta(seconds=round(seconds))), '%H:%M:%S').strftime('%Hh%Mm%Ss')
        return result


def merge_excel_files(filenames: list[str], output_filename: str):
    '''
    Merges the sheets of multiple excel files into a single excel file.
    If the target excel file already exists, appends without overwriting.
    
    ### Arguments
    #### Required
    - `filenames` (list[str]): list of filenames of Excel files to be merged.
    - `output_filename` (str): filename of the Excel file to be outputted.
    #### Optional
    '''    
    # init workbook
    sheet_names = set.intersection(*[set(pd.ExcelFile(filename).sheet_names) for filename in filenames])
    if not os.path.exists(output_filename):
        workbook = openpyxl.Workbook()
        workbook.save(output_filename)
        for sheet_name in sheet_names:
            workbook.create_sheet(title=sheet_name)
    else:
        workbook = openpyxl.load_workbook(output_filename)
    
    # merge the sheets
    for filename in filenames:
        with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            for sheet_name in sheet_names:
                sheet_exists = (sheet_name in writer.sheets)
                df = pd.read_excel(filename, sheet_name=sheet_name)
                start_i = writer.sheets[sheet_name].max_row if sheet_exists else 0
                df.to_excel(writer, sheet_name=sheet_name, index=False,
                        startrow=start_i, header=(not sheet_exists))
                print(f'Rows: {len(df)}, Start: {start_i}, Sheet: {sheet_name}, Source: {filename}, Dest: {output_filename}')


def reading_num_to_electrodes(
        index: int, as_str: bool = False) -> tuple[int, int, int, int, bool] | str:
    '''
    Converts a given index of the voltage output to a tuple of
    the electrode numbers being used to take the measurement.
    Assumes that the microcontroller is using the adjacent setting (AD).
    
    ### Arguments
    #### Required
    - `index` (int): index of the voltage output, between 0 and 1023 inclusive.
    #### Optional
    - `as_str` (bool, default = False): if True, show a readable string of the electrodes.
    
    ### Returns
    - `tuple[int, int, int, int, bool] | str`: (i_src, i_sink, v_pos, v_neg, is_zero).
    '''    
    assert 0 <= index <= 1023
    
    i_src = int(index // 32) + 1
    i_sink = (i_src + 1) % 32  # adjacent
    v_pos = index % 32 + 1
    v_neg = (v_pos + 1) % 32  # adjacent

    # the readings will be zero if (v_pos, v_neg) overlap with (src_pin, sink_pin)
    if len(set([v_pos, v_neg, i_src, i_sink])) != 4:
        is_zero = True
    else:
        is_zero = False

    if as_str:
        return f'I_in: {i_src}, I_out: {i_sink}, V+: {v_pos}, V-: {v_neg}, Is zeroes: {is_zero}'
    else:
        return i_src, i_sink, v_pos, v_neg, is_zero


def electrodes_to_reading_num(current_source: int, v_pos: int) -> int:
    '''
    Convert a desired electrode combination to the index of the voltage output.
    Note that when using the adjacent setting, the current sink is 1 more than
    the current source and the v_neg is 1 more than the v_pos (mod 32).
    
    ### Arguments
    #### Required
    - `current_source` (int): electrode number to input current at. 1 to 32 inclusive.
    - `v_pos` (int): electrode number to take as positive voltage. 1 to 32 inclusive.
    #### Optional
    
    ### Returns
    - `int`: index of corresponding voltage output, between 0 and 1023 inclusive.
    '''
    assert 1 <= current_source <= 32
    assert 1 <= v_pos <= 32
    return (current_source - 1) * 32 + (v_pos - 1)


def convert_xy_to_disp_map(positions: list | np.ndarray) -> np.ndarray:
    '''
    Given a set of frame coordinates, return an image-like array showing
    the position of the fingers. It is OK for some of the values to be None,
    which represent empty positions.
        
    ### Arguments
    #### Required
    - `positions` (list | np.ndarray): list of position coordinates
    [x0, x1, ..., y0, y1, ...], in metres in the frame coordinates.
    Shape: (num_trials, 2 * num_fingers) or can be (2 * num_fingers,) if only one.
    #### Optional
        
    ### Returns
    - `np.ndarray`: array of images showing the finger positions. Each array
    will be 1 if there is a press within that tile, and 0 otherwise.
    The radius of the finger is neglected.
    Shape: (num_trials, self.GRID_DIV_X, self.GRID_DIV_Y) or
    (self.GRID_DIV_X, self.GRID_DIV_Y) if only one.
    '''

    FRAME_X, FRAME_Y, GRID_DIV_X, GRID_DIV_Y = 0.08, 0.12, 6, 8

    tile_x = FRAME_X / GRID_DIV_X
    tile_y = FRAME_Y / GRID_DIV_Y

    if isinstance(positions, list):
        positions = np.array(positions)
    if positions.ndim == 1:
        positions = positions.reshape(1, -1)

    disp_maps = []
    for pos in positions:
        pos = pos[~np.isnan(pos)]  # remove None entries, preserve order
        disp_map = np.zeros((GRID_DIV_X, GRID_DIV_Y))
        x_list = pos[::len(pos) // 2]
        y_list = pos[len(pos) // 2::]
        for x, y in zip(x_list, y_list):
            disp_map[int(x // tile_x), int(y // tile_y)] = 1.0
        disp_maps.append(disp_map)
    return np.array(disp_maps) if len(disp_maps) > 1 else disp_maps[0]


def get_closest_single_points(X1: pd.DataFrame, multi_pos: np.ndarray | pd.DataFrame,
                              index_only: bool = True) -> list[int] | list[pd.DataFrame]:
    '''
    Given a position array of two fingers, finds the closest matching
    two points in the single finger dataset.
    
    ### Arguments
    #### Required
    - `X1` (pd.DataFrame): full single finger position dataset, without null columns.
    Shape: (N, 2), where N is the number
    - `pos_2` (np.ndarray): array of the form [x0, x1, ..., y0, y1, ...].
    Shape: (N, 2n), where N is the number of multi-touch data points and n is the number of fingers.
    #### Optional
    - `index_only` (bool, default = True): if True, return the indices of the
    matching rows, otherwise return the rows themselves.
    
    ### Returns
    - `list[int] | list[pd.DataFrame]`: matching indices or rows each in the form [[x0, y0], [x1, y1], ...]
    Shape: (N, n) for indices, or (N, n, 2) for rows.
    '''

    if isinstance(multi_pos, pd.DataFrame):
        multi_pos = multi_pos.values

    matches = []
    for pos_arr in multi_pos:
        matches_this_point = []
        for xi, yi in zip(range(int(len(pos_arr) // 2)), range(int(len(pos_arr) // 2), len(pos_arr))):
            pos = np.array([pos_arr[xi], pos_arr[yi]])
            match_idx = np.argmin(np.linalg.norm(X1 - pos, axis=1))
            if index_only:
                matches_this_point.append(match_idx)
            else:
                matches_this_point.append(X1.iloc[match_idx, :].values)
        matches.append(matches_this_point)
    return np.array(matches)


if __name__ == '__main__':
    #merge_excel_files(
    #    [f'output/EIT_Data_Gelatin_1_finger_2_dof_Set_{i}.xlsx' for i in range(1, 2)],
    #    'output/EIT_Data_Gelatin_1_finger_2_dof.xlsx'
    #)

    pos = pd.DataFrame([[0.05, None, 0.08, None], [0.05, 0.03, 0.08, 0.02]])
    print(pos)

    print(convert_xy_to_disp_map(pos.values))